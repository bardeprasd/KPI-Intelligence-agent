"""
Output packaging and export utilities for KPI results.

This file sits in the presentation/output layer of the project. It assembles
the final payload used by the API/chatbot layer and writes leadership-friendly
artifacts in JSON, Excel, HTML, and text formats.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Dict, List, Optional, Sequence

import pandas as pd
from jinja2 import Template
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import ASSUMPTIONS, CALCULATION_VERSION, SUMMARY_CARD_ORDER

# Color mapping used across spreadsheet and HTML status displays.
STATUS_FILL = {
    "green": "C6EFCE",
    "amber": "FFEB9C",
    "red": "FFC7CE",
    "info": "D9EAF7",
}


HTML_DRILLDOWN_TARGETS = {
    "Average Inbound Lead Time": "by_supplier",
    "Receipts On-Time %": "by_supplier",
    "Quantity Discrepancy %": "by_part_number",
    "Inbound Volume": "by_date",
    "Late Receipt Count": "by_supplier",
    "Top 5 Delaying Suppliers": "by_late_supplier_rank",
    "Fill Rate %": "by_part_number",
    "OTIF %": "by_customer",
    "Backorder Rate %": "by_part_number",
    "Outbound Volume": "by_date",
    "Late Shipment Count": "by_customer",
    "Top 10 SKUs by Backorder": "by_backorder_sku_rank",
    "Days of Supply": "by_sku_family",
    "Stockout Exposure %": "by_sku_family",
    "Safety Stock Coverage %": "by_warehouse",
    "Aged Inventory % (>180d)": "by_part_number",
    "Average Inventory Age": "by_part_number",
    "Lines Picked per Labor-Hour": "by_warehouse",
    "Orders Processed per Labor-Hour": "by_warehouse",
    "Orders per Day": "by_date",
    "SLA Adherence %": "by_warehouse",
    "Equipment Utilization %": "by_warehouse",
    "Touches per Order": "by_warehouse",
    "Picks per Person per Hour": "by_employee",
    "Error Rate %": "by_employee",
    "Rework Rate %": "by_employee",
    "Overtime %": "by_employee",
    "Average Tasks per Employee": "by_employee",
}

GROUP_RAW_DETAIL_COLUMN_MAP = {
    "receipt_date_day": "received_date",
    "ship_date_day": "shipped_date",
    "snapshot_date_day": "snapshot_date",
    "operation_date_day": "date",
    "work_date_day": "date",
}


def _html_slug(text: str) -> str:
    return "".join(ch.lower() if ch.isalnum() else "-" for ch in text).strip("-")


def _drilldown_anchor(section_name: str, drilldown_id: str, grain: Optional[str] = None) -> str:
    anchor = f"drilldown-{_html_slug(section_name)}-{_html_slug(drilldown_id)}"
    if grain:
        anchor = f"{anchor}-{_html_slug(grain)}"
    return anchor


def _kpi_drilldown_anchor(section_name: str, kpi_name: str) -> str:
    return f"kpi-drilldown-{_html_slug(section_name)}-{_html_slug(kpi_name)}"


def _primary_drilldown_anchor(section: Dict, kpi_name: Optional[str] = None) -> Optional[str]:
    drilldowns = section.get("drilldowns", {})
    preferred = HTML_DRILLDOWN_TARGETS.get(kpi_name or "")
    candidates = [preferred] if preferred else []
    candidates.extend(drilldowns.keys())
    seen = set()
    for drilldown_id in candidates:
        if not drilldown_id or drilldown_id in seen:
            continue
        seen.add(drilldown_id)
        drilldown = drilldowns.get(drilldown_id)
        if not drilldown:
            continue
        if drilldown.get("grain_tables"):
            return _drilldown_anchor(section["name"], drilldown_id, "day")
        return _drilldown_anchor(section["name"], drilldown_id)
    return None


def _preferred_drilldown(section: Dict, kpi_name: str) -> Optional[Dict]:
    drilldown_id = HTML_DRILLDOWN_TARGETS.get(kpi_name)
    if not drilldown_id:
        return None
    drilldown = section.get("drilldowns", {}).get(drilldown_id)
    if not drilldown:
        return None
    if drilldown.get("grain_tables"):
        grain_table = drilldown["grain_tables"].get("day")
        if grain_table:
            return {
                "drilldown_id": drilldown_id,
                "label": f"{drilldown.get('label', drilldown_id)} (Day)",
                "table": grain_table,
                "base_drilldown": drilldown,
            }
    return {
        "drilldown_id": drilldown_id,
        "label": drilldown.get("label", drilldown_id),
        "table": drilldown,
        "base_drilldown": drilldown,
    }


def _normalize_group_match_value(value: object) -> Optional[str]:
    if value is None:
        return None
    return str(value)


def _group_matches_raw_row(raw_row: Dict, group_dimensions: Dict[str, object]) -> bool:
    for group_column, group_value in group_dimensions.items():
        if group_column == "rank":
            continue
        raw_column = GROUP_RAW_DETAIL_COLUMN_MAP.get(group_column, group_column)
        raw_value = raw_row.get(raw_column)
        if _normalize_group_match_value(raw_value) != _normalize_group_match_value(group_value):
            return False
    return True


def _build_group_raw_detail(raw_detail: Optional[Dict], group_dimensions: Dict[str, object]) -> Optional[Dict]:
    if not raw_detail:
        return None
    matching_rows = [row for row in raw_detail.get("rows", []) if _group_matches_raw_row(row, group_dimensions)]
    return {
        "columns": raw_detail.get("columns", []),
        "row_count": len(matching_rows),
        "logic_note": "Raw records filtered to the selected grouped row.",
        "rows": matching_rows,
    }


def _build_html_kpi_drilldown(section: Dict, kpi: Dict) -> Optional[Dict]:
    preferred = _preferred_drilldown(section, kpi["name"])
    if not preferred:
        return None

    drilldown = preferred["base_drilldown"]
    table = preferred["table"]
    payload = {
        "anchor": _kpi_drilldown_anchor(section["name"], kpi["name"]),
        "section_name": section["name"],
        "kpi_name": kpi["name"],
        "label": preferred["label"],
        "source_dataset": drilldown.get("source_dataset", ""),
        "applied_filters": drilldown.get("applied_filters", {}),
        "logic_note": drilldown.get("logic_note", ""),
        "group_by": [column for column in table.get("group_by", []) if column != "rank"],
        "dimension_labels": table.get("dimension_labels", {}),
        "available": bool(drilldown.get("available")),
        "rows": [],
        "unavailable_reason": drilldown.get("unavailable_reason"),
    }

    if not drilldown.get("available"):
        return payload

    rows: List[Dict] = []
    for row in table.get("rows", []):
        metric = row.get("metrics", {}).get(kpi["name"])
        if not metric:
            metric_values = list(row.get("metrics", {}).values())
            if len(metric_values) == 1:
                metric = metric_values[0]
        if not metric:
            continue
        group_dimensions = {column: row.get(column) for column in table.get("group_by", [])}
        rows.append({
            "dimensions": {column: row.get(column) for column in table.get("group_by", [])},
            "row_count": row.get("row_count"),
            "rank": row.get("rank"),
            "metric": metric,
            "group_anchor": f"{payload['anchor']}-group-{len(rows) + 1}",
            "raw_detail": _build_group_raw_detail(kpi.get("raw_detail"), group_dimensions),
        })
    payload["rows"] = rows
    if not rows and payload["available"]:
        payload["unavailable_reason"] = f"No drill-down rows were available for {kpi['name']}."
    return payload


def _flatten_drilldown_rows(section: Dict) -> List[Dict]:
    rows: List[Dict] = []
    for drilldown_id, drilldown in section.get("drilldowns", {}).items():
        if not drilldown.get("available"):
            rows.append({
                "section": section["name"],
                "drilldown": drilldown_id,
                "grain": None,
                "available": False,
                "reason": drilldown.get("unavailable_reason"),
                "dimension_values": "",
                "metric": "",
                "display_value": "",
                "status": "",
                "formula": "",
                "source_dataset": drilldown.get("source_dataset", ""),
            })
            continue

        grain_tables = drilldown.get("grain_tables")
        if grain_tables:
            for grain, grain_table in grain_tables.items():
                rows.extend(_flatten_single_drilldown(section["name"], drilldown_id, grain_table, grain))
        else:
            rows.extend(_flatten_single_drilldown(section["name"], drilldown_id, drilldown, None))
    return rows


def _flatten_single_drilldown(section_name: str, drilldown_id: str, drilldown: Dict, grain: Optional[str]) -> List[Dict]:
    flattened: List[Dict] = []
    group_by = drilldown.get("group_by", [])
    for row in drilldown.get("rows", []):
        dimensions = ", ".join(f"{column}={row.get(column)}" for column in group_by)
        for metric_name, metric in row.get("metrics", {}).items():
            flattened.append({
                "section": section_name,
                "drilldown": drilldown_id,
                "grain": grain,
                "available": True,
                "reason": None,
                "dimension_values": dimensions,
                "metric": metric_name,
                "display_value": metric.get("display_value", ""),
                "status": metric.get("status", ""),
                "formula": metric.get("formula", ""),
                "source_dataset": drilldown.get("source_dataset", ""),
            })
    return flattened


def overall_status(summary_cards: List[Dict]) -> str:
    # ================================
    # Function: overall_status
    # Purpose: Collapses summary-card statuses into one executive health signal.
    # Inputs:
    #   - summary_cards (List[Dict]): top KPI cards for the report
    # Output:
    #   - str overall status: red, amber, or green
    # ================================
    statuses = [card["status"] for card in summary_cards]
    if "red" in statuses:
        return "red"
    if "amber" in statuses:
        return "amber"
    return "green"


def build_summary_cards(sections: List[Dict]) -> List[Dict]:
    # ================================
    # Function: build_summary_cards
    # Purpose: Selects the KPI cards shown in the report header.
    # Inputs:
    #   - sections (List[Dict]): KPI sections
    # Output:
    #   - List[Dict] containing summary-card payloads in configured order
    # ================================
    lookup = {kpi["name"]: (section, kpi) for section in sections for kpi in section["kpis"]}
    cards = []
    for name in SUMMARY_CARD_ORDER:
        section, k = lookup[name]
        cards.append({
            "label": k["name"],
            "value": k["value"],
            "display_value": k["display_value"],
            "unit": k["unit"],
            "status": k["status"],
            "target": k["target"],
            "target_display": k["target_display"],
            "note": k.get("note", ""),
            "drilldown_anchor": _kpi_drilldown_anchor(section["name"], k["name"]),
        })
    return cards


def build_payload(
    sections: List[Dict],
    kpi_table: List[Dict],
    insights: List[str],
    risks: List[str],
    recommendations: List[str],
    row_counts: Dict[str, int],
    validation_warnings: List[str],
    start_date: pd.Timestamp,
    end_date: pd.Timestamp,
    period_label: str,
    warehouse_filter: Optional[Sequence[str]],
    sku_family_filter: Optional[Sequence[str]] = None,
    llm_used: bool = False,
):
    # ================================
    # Function: build_payload
    # Purpose: Packages all KPI results and metadata into one canonical payload.
    # Inputs:
    #   - sections, kpi_table, insights, risks, recommendations
    #   - row_counts, validation_warnings
    #   - start_date, end_date, period_label
    #   - warehouse_filter, sku_family_filter
    # Output:
    #   - Dict payload consumed by exporters and the chatbot
    # Important Logic:
    #   - Stores traceability, assumptions, and filter scope alongside KPI data
    # ================================
    summary_cards = build_summary_cards(sections)
    section_lookup = {section["name"]: section for section in sections}
    kpi_anchor_lookup: Dict[tuple[str, str], Optional[str]] = {}
    for section in sections:
        section["drilldown_anchor"] = _primary_drilldown_anchor(section)
        section["html_kpi_drilldowns"] = []
        for drilldown_id, drilldown in section.get("drilldowns", {}).items():
            drilldown["html_anchor"] = _drilldown_anchor(section["name"], drilldown_id)
            if drilldown.get("grain_tables"):
                for grain, grain_table in drilldown["grain_tables"].items():
                    grain_table["html_anchor"] = _drilldown_anchor(section["name"], drilldown_id, grain)
        for kpi in section["kpis"]:
            kpi["drilldown_anchor"] = None
            html_drilldown = _build_html_kpi_drilldown(section, kpi)
            if html_drilldown:
                kpi["drilldown_anchor"] = html_drilldown["anchor"]
                section["html_kpi_drilldowns"].append(html_drilldown)
            kpi_anchor_lookup[(section["name"], kpi["name"])] = kpi["drilldown_anchor"]
    for card in summary_cards:
        section = section_lookup.get(next(row["domain"] for row in kpi_table if row["kpi"] == card["label"]))
        card["drilldown_anchor"] = kpi_anchor_lookup.get((section["name"], card["label"])) if section else None
    for row in kpi_table:
        row["drilldown_anchor"] = kpi_anchor_lookup.get((row["domain"], row["kpi"]))
    payload = {
        "header": {
            "title": "Aftersales Operations KPI Intelligence Summary",
            "subtitle": "Executive view across inbound, outbound, inventory, warehouse, and workforce performance",
            "generated_at": pd.Timestamp.utcnow().isoformat(),
            "overall_status": overall_status(summary_cards),
            "warehouse_scope": list(warehouse_filter) if warehouse_filter else ["All"],
            "sku_family_scope": list(sku_family_filter) if sku_family_filter else ["All"],
            "calculation_version": CALCULATION_VERSION,
        },
        "reporting_period": {
            "start_date": start_date.date().isoformat(),
            "end_date": end_date.date().isoformat(),
            "label": period_label,
        },
        "summary_cards": summary_cards,
        "sections": sections,
        "kpi_table": kpi_table,
        "insights": insights,
        "risks": risks,
        "recommendations": recommendations,
        "metadata": {
            "source_files": [
                "inbound_parts.csv",
                "outbound_parts.csv",
                "inventory_snapshot.csv",
                "warehouse_productivity.csv",
                "employee_productivity.csv",
            ],
            "row_counts": row_counts,
            "warehouse_filter": list(warehouse_filter) if warehouse_filter else None,
            "sku_family_filter": list(sku_family_filter) if sku_family_filter else None,
            "calculation_version": CALCULATION_VERSION,
            "narrative_generation": "llm_refined" if llm_used else "deterministic",
            "risk_generation": "deterministic",
            "drilldown_enabled": True,
            "generated_files": [],
        },
        "audit": {
            "assumptions": ASSUMPTIONS,
            "validation_warnings": validation_warnings,
            "formula_traceability": True,
            "period_logic": "Latest full common month across operational event dates when start/end are not explicitly supplied.",
            "status_logic": "KPI statuses are assigned from config-driven threshold rules (green/amber/red/info).",
            "drilldown_traceability": "Each drill-down table includes source dataset, applied filters, and deterministic formula notes.",
        },
    }
    return payload


def build_assignment_contract(payload: Dict) -> Dict:
    # ================================
    # Function: build_assignment_contract
    # Purpose: Maps the rich internal payload to the strict assignment schema.
    # Inputs:
    #   - payload (Dict): rich KPI payload
    # Output:
    #   - Dict matching the published assignment contract
    # ================================
    summary_cards = []
    for card in payload["summary_cards"]:
        target = float(card["target"]) if card.get("target") is not None else 0.0
        delta = round(float(card["value"]) - target, 4) if card.get("target") is not None else None
        summary_cards.append({
            "name": card["label"],
            "current": card["display_value"],
            "target": target,
            "delta": delta,
            "status": card["status"],
            "comment": card.get("note") or None,
        })

    return {
        "header": {
            "period": payload["reporting_period"]["label"],
            "warehouses": payload["header"]["warehouse_scope"],
            "status": payload["header"]["overall_status"],
        },
        "summary_cards": summary_cards,
    }


def _validate_assignment_contract(contract: Dict) -> None:
    # ================================
    # Function: _validate_assignment_contract
    # Purpose: Validates the exported assignment JSON against the local schema.
    # Inputs:
    #   - contract (Dict): assignment-format payload
    # Output:
    #   - None
    # ================================
    if not isinstance(contract, dict):
        raise ValueError("Assignment output must be an object.")
    if not isinstance(contract.get("header"), dict):
        raise ValueError("Assignment output missing object field: header")
    if not isinstance(contract.get("summary_cards"), list):
        raise ValueError("Assignment output missing array field: summary_cards")

    header = contract["header"]
    for field in ["period", "warehouses", "status"]:
        if field not in header:
            raise ValueError(f"Assignment output missing header field: {field}")
    if not isinstance(header["period"], str):
        raise ValueError("Assignment header.period must be a string")
    if not isinstance(header["warehouses"], list) or not all(isinstance(item, str) for item in header["warehouses"]):
        raise ValueError("Assignment header.warehouses must be an array of strings")
    if not isinstance(header["status"], str):
        raise ValueError("Assignment header.status must be a string")

    for idx, card in enumerate(contract["summary_cards"]):
        if not isinstance(card, dict):
            raise ValueError(f"summary_cards[{idx}] must be an object")
        for field in ["name", "current", "target", "delta", "status", "comment"]:
            if field not in card:
                raise ValueError(f"summary_cards[{idx}] missing field: {field}")
        if not isinstance(card["name"], str):
            raise ValueError(f"summary_cards[{idx}].name must be a string")
        if not isinstance(card["current"], (str, int, float)):
            raise ValueError(f"summary_cards[{idx}].current must be a string or number")
        if not isinstance(card["target"], (int, float)):
            raise ValueError(f"summary_cards[{idx}].target must be a number")
        if card["delta"] is not None and not isinstance(card["delta"], (int, float)):
            raise ValueError(f"summary_cards[{idx}].delta must be a number or null")
        if not isinstance(card["status"], str):
            raise ValueError(f"summary_cards[{idx}].status must be a string")
        if card["comment"] is not None and not isinstance(card["comment"], str):
            raise ValueError(f"summary_cards[{idx}].comment must be a string or null")


def write_json(payload: Dict, output_dir: Path) -> Path:
    # ================================
    # Function: write_json
    # Purpose: Writes the strict assignment JSON plus a richer detailed JSON.
    # Inputs:
    #   - payload (Dict): packaged KPI output
    #   - output_dir (Path): target directory
    # Output:
    #   - Path to the written JSON file
    # ================================
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / "leadership_summary.json"
    detailed_path = output_dir / "leadership_summary_detailed.json"
    assignment_contract = build_assignment_contract(payload)
    _validate_assignment_contract(assignment_contract)
    path.write_text(json.dumps(assignment_contract, indent=2), encoding="utf-8")
    detailed_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return path


def _apply_status_fill(cell, status: str):
    # ================================
    # Function: _apply_status_fill
    # Purpose: Applies traffic-light coloring to a spreadsheet cell.
    # Inputs:
    #   - cell: openpyxl cell object
    #   - status (str): KPI status key
    # Output:
    #   - None
    # ================================
    cell.fill = PatternFill(fill_type="solid", start_color=STATUS_FILL.get(status, "FFFFFF"), end_color=STATUS_FILL.get(status, "FFFFFF"))


def write_excel(payload: Dict, output_dir: Path) -> Path:
    # ================================
    # Function: write_excel
    # Purpose: Writes a multi-sheet Excel workbook for leadership review.
    # Inputs:
    #   - payload (Dict): packaged KPI output
    #   - output_dir (Path): target directory
    # Output:
    #   - Path to the written Excel workbook
    # Important Logic:
    #   - Splits high-level KPI, section detail, insights, and metadata into
    #     separate sheets for easier stakeholder consumption
    # ================================
    from datetime import datetime

    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / "kpi_summary.xlsx"

    def _timestamped_path() -> Path:
        return output_dir / f"kpi_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    def _is_file_locked(target: Path) -> bool:
        if not target.exists():
            return False
        try:
            with open(target, "a+b"):
                return False
        except PermissionError:
            return True
        except OSError:
            return True

    wb = Workbook()
    ws = wb.active
    ws.title = "KPI_Summary"
    headers = ["Domain", "KPI", "Current", "Target", "Delta", "Status", "Comment"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in payload["kpi_table"]:
        delta = round(float(row["value"]) - float(row["target"]), 4) if row.get("target") is not None else None
        ws.append([
            row["domain"],
            row["kpi"],
            row["display_value"],
            row["target_display"],
            delta,
            row["status"].upper(),
            row["trend_note"],
        ])
        _apply_status_fill(ws.cell(ws.max_row, 6), row["status"])

    detail = wb.create_sheet("Section_Detail")
    detail.append(["Section", "Risk Level", "Insight", "KPI", "Value", "Target", "Status", "Formula", "Source"])
    for cell in detail[1]:
        cell.font = Font(bold=True)
    for section in payload["sections"]:
        for kpi in section["kpis"]:
            detail.append([
                section["name"], section["risk_level"].upper(), section["insight"], kpi["name"], kpi["display_value"],
                kpi["target_display"], kpi["status"].upper(), kpi["formula"], kpi["source_table"]
            ])
            _apply_status_fill(detail.cell(detail.max_row, 7), kpi["status"])

    drilldown_ws = wb.create_sheet("Drilldowns")
    drilldown_ws.append(["Section", "Drilldown", "Time Grain", "Dimension Values", "Metric", "Value", "Status", "Formula", "Source Dataset", "Availability / Reason"])
    for cell in drilldown_ws[1]:
        cell.font = Font(bold=True)
    for section in payload["sections"]:
        for row in _flatten_drilldown_rows(section):
            drilldown_ws.append([
                row["section"],
                row["drilldown"],
                row["grain"] or "",
                row["dimension_values"],
                row["metric"],
                row["display_value"],
                row["status"].upper() if row["status"] else "",
                row["formula"],
                row["source_dataset"],
                "AVAILABLE" if row["available"] else (row["reason"] or "UNAVAILABLE"),
            ])
            if row["status"]:
                _apply_status_fill(drilldown_ws.cell(drilldown_ws.max_row, 7), row["status"])

    insights_ws = wb.create_sheet("Insights_Actions")
    insights_ws.append(["Type", "Text"])
    for cell in insights_ws[1]:
        cell.font = Font(bold=True)
    for insight in payload["insights"]:
        insights_ws.append(["Insight", insight])
    for risk in payload.get("risks", []):
        insights_ws.append(["Risk", risk])
    for recommendation in payload["recommendations"]:
        insights_ws.append(["Recommendation", recommendation])

    meta = wb.create_sheet("Metadata")
    meta.append(["Field", "Value"])
    for cell in meta[1]:
        cell.font = Font(bold=True)
    meta_rows = [
        ("Reporting Period", payload["reporting_period"]["label"]),
        ("Overall Status", payload["header"]["overall_status"].upper()),
        ("Warehouse Scope", ", ".join(payload["header"]["warehouse_scope"])),
        ("Calculation Version", payload["header"]["calculation_version"]),
        ("Validation Warning Count", len(payload["audit"]["validation_warnings"])),
    ]
    meta_rows.extend((f"Row Count - {k}", v) for k, v in payload["metadata"]["row_counts"].items())
    for item in payload["audit"]["validation_warnings"]:
        meta_rows.append(("Validation Warning", item))
    for row in meta_rows:
        meta.append(list(row))

    # Auto-size columns to keep the workbook readable without manual cleanup.
    for ws_obj in wb.worksheets:
        for col_idx in range(1, ws_obj.max_column + 1):
            length = 0
            for row_idx in range(1, ws_obj.max_row + 1):
                value = ws_obj.cell(row_idx, col_idx).value
                length = max(length, len(str(value)) if value is not None else 0)
            ws_obj.column_dimensions[get_column_letter(col_idx)].width = min(length + 2, 60)

    if _is_file_locked(path):
        path = _timestamped_path()

    try:
        wb.save(path)
    except PermissionError:
        if path.name == "kpi_summary.xlsx":
            fallback_path = _timestamped_path()
            wb.save(fallback_path)
            path = fallback_path
        else:
            raise
    return path


HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>{{ payload.header.title }}</title>
  <style>
    :root {
      --bg: #eef2f6;
      --bg-deep: #e3eaf1;
      --surface-strong: #ffffff;
      --text: #142233;
      --text-strong: #0c1726;
      --muted: #617186;
      --muted-strong: #445468;
      --border: rgba(19, 35, 56, 0.1);
      --accent: #274c77;
      --green: #26775a;
      --green-soft: rgba(38, 119, 90, 0.12);
      --amber: #a66a17;
      --amber-soft: rgba(166, 106, 23, 0.12);
      --red: #b23a31;
      --red-soft: rgba(178, 58, 49, 0.12);
      --info: #37648d;
      --info-soft: rgba(55, 100, 141, 0.12);
      --hero-start: #122438;
      --hero-end: #1b3651;
      --radius-lg: 28px;
      --space-3: 16px;
      --space-4: 22px;
      --space-5: 28px;
      --space-6: 36px;
      --shadow-sm: 0 8px 22px rgba(13, 26, 42, 0.06);
      --shadow-md: 0 18px 44px rgba(13, 26, 42, 0.1);
      --shadow-lg: 0 28px 64px rgba(10, 18, 31, 0.14);
      --font-ui: "Aptos", "Segoe UI Variable Text", "Segoe UI", "Helvetica Neue", Arial, sans-serif;
    }

    * { box-sizing: border-box; }
    html { color-scheme: light; }

    body {
      margin: 0;
      min-width: 320px;
      background:
        radial-gradient(circle at top left, rgba(255, 255, 255, 0.92) 0%, rgba(255, 255, 255, 0) 34%),
        linear-gradient(180deg, #f5f7fa 0%, #eef2f6 38%, #e3eaf1 100%);
      color: var(--text);
      font-family: var(--font-ui);
      -webkit-font-smoothing: antialiased;
      text-rendering: optimizeLegibility;
    }

    body::before {
      content: "";
      position: fixed;
      inset: 0;
      pointer-events: none;
      background:
        linear-gradient(rgba(255, 255, 255, 0.42), rgba(255, 255, 255, 0.42)),
        linear-gradient(90deg, rgba(19, 35, 56, 0.02) 1px, transparent 1px),
        linear-gradient(rgba(19, 35, 56, 0.02) 1px, transparent 1px);
      background-size: auto, 72px 72px, 72px 72px;
      mask-image: linear-gradient(180deg, rgba(0, 0, 0, 0.24), transparent 85%);
    }

    .page {
      width: min(1480px, calc(100vw - 64px));
      margin: 12px auto 28px;
      display: grid;
      gap: var(--space-5);
      position: relative;
      z-index: 1;
    }

    .hero {
      position: relative;
      overflow: hidden;
      border-radius: var(--radius-lg);
      padding: var(--space-6);
      color: #f5f8fc;
      background:
        radial-gradient(circle at 85% 22%, rgba(128, 183, 255, 0.24), rgba(128, 183, 255, 0) 26%),
        radial-gradient(circle at 20% 120%, rgba(255, 255, 255, 0.12), rgba(255, 255, 255, 0) 24%),
        linear-gradient(135deg, var(--hero-start) 0%, var(--hero-end) 62%, #244969 100%);
      box-shadow: var(--shadow-lg);
      border: 1px solid rgba(255, 255, 255, 0.08);
      isolation: isolate;
    }

    .hero::before,
    .hero::after {
      content: "";
      position: absolute;
      border-radius: 999px;
      pointer-events: none;
    }

    .hero::before {
      width: 420px;
      height: 420px;
      right: -120px;
      top: -160px;
      background: radial-gradient(circle, rgba(255, 255, 255, 0.18) 0%, rgba(255, 255, 255, 0) 68%);
    }

    .hero::after {
      width: 520px;
      height: 520px;
      left: -240px;
      bottom: -360px;
      background: radial-gradient(circle, rgba(102, 151, 205, 0.22) 0%, rgba(102, 151, 205, 0) 72%);
    }

    .hero-top,
    .hero-metadata {
      position: relative;
      z-index: 1;
    }

    .hero-top {
      display: grid;
      grid-template-columns: minmax(0, 1fr) auto;
      gap: var(--space-4);
      align-items: start;
    }

    .hero-copy { max-width: 1120px; }

    .eyebrow {
      display: inline-flex;
      align-items: center;
      gap: 10px;
      margin-bottom: 14px;
      font-size: 12px;
      font-weight: 700;
      letter-spacing: 0.18em;
      text-transform: uppercase;
      color: rgba(245, 248, 252, 0.8);
    }

    .eyebrow::before {
      content: "";
      width: 42px;
      height: 1px;
      background: rgba(255, 255, 255, 0.45);
    }

    h1 {
      margin: 0;
      font-size: clamp(38px, 4.4vw, 64px);
      line-height: 0.98;
      letter-spacing: -0.04em;
      font-weight: 800;
      color: #ffffff;
    }

    .subtitle {
      margin-top: 16px;
      max-width: 980px;
      color: rgba(245, 248, 252, 0.78);
      font-size: 19px;
      line-height: 1.55;
    }

    .status-badge {
      display: inline-flex;
      align-items: center;
      gap: 10px;
      border-radius: 999px;
      padding: 14px 18px;
      font-size: 12px;
      font-weight: 800;
      letter-spacing: 0.12em;
      text-transform: uppercase;
      border: 1px solid rgba(255, 255, 255, 0.16);
      background: rgba(255, 255, 255, 0.08);
      color: #ffffff;
      box-shadow: inset 0 1px 0 rgba(255, 255, 255, 0.08);
      white-space: nowrap;
      backdrop-filter: blur(14px);
    }

    .status-badge::before {
      content: "";
      width: 9px;
      height: 9px;
      border-radius: 50%;
      background: currentColor;
      box-shadow: 0 0 0 4px rgba(255, 255, 255, 0.08);
    }

    .status-badge.red { color: #ffb4ad; }
    .status-badge.amber { color: #ffd597; }
    .status-badge.green { color: #aae4ca; }
    .status-badge.info { color: #b9d8f1; }

    .hero-metadata {
      display: grid;
      grid-template-columns: repeat(5, minmax(220px, 1fr));
      gap: var(--space-3);
      margin-top: var(--space-5);
    }

    .meta-tile {
      min-height: 122px;
      padding: 18px 18px 20px;
      border-radius: 20px;
      background: rgba(255, 255, 255, 0.08);
      border: 1px solid rgba(255, 255, 255, 0.12);
      box-shadow: inset 0 1px 0 rgba(255, 255, 255, 0.06);
      backdrop-filter: blur(14px);
    }

    .meta-label {
      font-size: 11px;
      letter-spacing: 0.14em;
      text-transform: uppercase;
      font-weight: 700;
      color: rgba(245, 248, 252, 0.62);
      margin-bottom: 12px;
    }

    .meta-value {
      font-size: 17px;
      line-height: 1.5;
      font-weight: 700;
      color: #ffffff;
      word-break: break-word;
    }

    .summary-grid {
      display: grid;
      grid-template-columns: repeat(5, minmax(240px, 1fr));
      gap: var(--space-4);
    }

    .summary-card,
    .panel,
    .table-panel {
      background: linear-gradient(180deg, rgba(255, 255, 255, 0.95) 0%, var(--surface-strong) 100%);
      border: 1px solid var(--border);
      box-shadow: var(--shadow-md);
    }

    .summary-card {
      position: relative;
      overflow: hidden;
      min-height: 224px;
      padding: 22px 20px;
      border-radius: 22px;
      display: grid;
      grid-template-rows: auto auto 1fr;
      gap: var(--space-3);
      transition: transform 180ms ease, box-shadow 180ms ease;
    }

    .summary-card::before {
      content: "";
      position: absolute;
      inset: 0 0 auto;
      height: 5px;
      background: linear-gradient(90deg, var(--accent) 0%, rgba(39, 76, 119, 0.45) 100%);
    }

    .summary-card::after {
      content: "";
      position: absolute;
      inset: auto -28px -44px auto;
      width: 140px;
      height: 140px;
      border-radius: 50%;
      background: radial-gradient(circle, rgba(39, 76, 119, 0.08) 0%, rgba(39, 76, 119, 0) 68%);
      pointer-events: none;
    }

    .summary-card.red::before { background: linear-gradient(90deg, var(--red) 0%, rgba(178, 58, 49, 0.35) 100%); }
    .summary-card.amber::before { background: linear-gradient(90deg, var(--amber) 0%, rgba(166, 106, 23, 0.35) 100%); }
    .summary-card.green::before { background: linear-gradient(90deg, var(--green) 0%, rgba(38, 119, 90, 0.35) 100%); }
    .summary-card.info::before { background: linear-gradient(90deg, var(--info) 0%, rgba(55, 100, 141, 0.35) 100%); }
    .summary-card:hover { transform: translateY(-2px); box-shadow: var(--shadow-lg); }

    .drilldown-link {
      display: inline-flex;
      align-items: center;
      gap: 8px;
      margin-top: 10px;
      font-size: 11px;
      font-weight: 800;
      letter-spacing: 0.08em;
      text-transform: uppercase;
      color: var(--accent);
      text-decoration: none;
    }

    .drilldown-link:hover { text-decoration: underline; }
    .drilldown-link::before { content: "↓"; font-size: 12px; line-height: 1; }

    .drilldown-jump {
      display: inline-flex;
      align-items: center;
      gap: 6px;
      margin-top: 8px;
      font-size: 10px;
      font-weight: 800;
      letter-spacing: 0.08em;
      text-transform: uppercase;
      color: var(--accent);
      text-decoration: none;
    }

    .drilldown-jump:hover { text-decoration: underline; }
    .drilldown-jump::before { content: "↓"; font-size: 11px; line-height: 1; }

    .card-top {
      display: flex;
      justify-content: space-between;
      gap: var(--space-3);
      align-items: flex-start;
    }

    .card-label {
      font-size: 13px;
      line-height: 1.45;
      letter-spacing: 0.03em;
      color: var(--muted-strong);
      font-weight: 800;
      text-transform: uppercase;
      max-width: 78%;
    }

    .pill {
      display: inline-flex;
      align-items: center;
      justify-content: center;
      gap: 6px;
      border-radius: 999px;
      padding: 6px 10px;
      font-size: 10px;
      line-height: 1;
      letter-spacing: 0.11em;
      font-weight: 800;
      text-transform: uppercase;
      border: 1px solid transparent;
      white-space: nowrap;
    }

    .pill.red { background: var(--red-soft); color: var(--red); border-color: rgba(178, 58, 49, 0.18); }
    .pill.amber { background: var(--amber-soft); color: var(--amber); border-color: rgba(166, 106, 23, 0.18); }
    .pill.green { background: var(--green-soft); color: var(--green); border-color: rgba(38, 119, 90, 0.18); }
    .pill.info { background: var(--info-soft); color: var(--info); border-color: rgba(55, 100, 141, 0.18); }

    .card-value {
      font-size: clamp(42px, 3.4vw, 56px);
      line-height: 1;
      letter-spacing: -0.05em;
      font-weight: 800;
      color: var(--text-strong);
    }

    .card-footer {
      align-self: end;
      display: grid;
      gap: 10px;
      padding-top: 4px;
      font-size: 13px;
      line-height: 1.5;
      color: var(--muted);
    }

    .card-target,
    .card-note {
      padding-top: 10px;
      border-top: 1px solid rgba(19, 35, 56, 0.08);
    }

    .card-target strong,
    .card-note strong { color: var(--text); }

    .main-grid {
      display: grid;
      grid-template-columns: minmax(0, 1.9fr) minmax(420px, 0.95fr);
      gap: var(--space-5);
      align-items: start;
    }

    .panel,
    .table-panel {
      border-radius: 24px;
      overflow: hidden;
      backdrop-filter: blur(10px);
    }

    .panel-header {
      display: flex;
      justify-content: space-between;
      align-items: flex-start;
      gap: var(--space-3);
      padding: var(--space-4) var(--space-5);
      border-bottom: 1px solid rgba(19, 35, 56, 0.08);
      background: linear-gradient(180deg, rgba(246, 249, 252, 0.95) 0%, rgba(255, 255, 255, 0.75) 100%);
    }

    .panel-title {
      margin: 0;
      font-size: 20px;
      line-height: 1.2;
      font-weight: 800;
      letter-spacing: -0.02em;
      color: var(--text-strong);
    }

    .panel-subtitle {
      margin-top: 6px;
      font-size: 14px;
      line-height: 1.5;
      color: var(--muted);
      max-width: 720px;
    }

    .section-grid {
      display: grid;
      grid-template-columns: repeat(2, minmax(360px, 1fr));
      gap: var(--space-4);
      padding: var(--space-4);
      background: linear-gradient(180deg, rgba(248, 250, 252, 0.6) 0%, rgba(255, 255, 255, 0) 100%);
    }

    .section-card {
      position: relative;
      min-height: 328px;
      padding: 22px;
      border-radius: 22px;
      border: 1px solid rgba(19, 35, 56, 0.08);
      background: linear-gradient(180deg, rgba(255, 255, 255, 0.98) 0%, rgba(248, 250, 252, 0.92) 100%);
      box-shadow: var(--shadow-sm);
      overflow: hidden;
    }

    .section-card::before {
      content: "";
      position: absolute;
      inset: 0 auto 0 0;
      width: 5px;
      background: var(--accent);
    }

    .section-card::after {
      content: "";
      position: absolute;
      right: -26px;
      top: -26px;
      width: 110px;
      height: 110px;
      border-radius: 50%;
      background: radial-gradient(circle, rgba(39, 76, 119, 0.08) 0%, rgba(39, 76, 119, 0) 72%);
      pointer-events: none;
    }

    .section-card.red::before { background: var(--red); }
    .section-card.amber::before { background: var(--amber); }
    .section-card.green::before { background: var(--green); }
    .section-card.info::before { background: var(--info); }

    .section-top {
      display: flex;
      justify-content: space-between;
      gap: var(--space-3);
      align-items: flex-start;
      margin-bottom: var(--space-3);
      padding-left: 6px;
    }

    .section-title {
      margin: 0;
      font-size: 20px;
      line-height: 1.2;
      font-weight: 800;
      color: var(--text-strong);
      letter-spacing: -0.02em;
    }

    .metric-list {
      list-style: none;
      padding: 0;
      margin: 0;
      display: grid;
      gap: 12px;
    }

    .metric-row {
      display: grid;
      grid-template-columns: 14px minmax(0, 1fr) auto;
      gap: 14px;
      align-items: center;
      padding: 13px 14px;
      border-radius: 16px;
      background: rgba(246, 249, 252, 0.78);
      border: 1px solid rgba(19, 35, 56, 0.06);
      font-size: 13px;
    }

    .metric-name {
      color: var(--muted-strong);
      line-height: 1.5;
      font-weight: 600;
      min-width: 0;
      overflow-wrap: anywhere;
    }

    .metric-value {
      font-weight: 800;
      color: var(--text-strong);
      white-space: nowrap;
      letter-spacing: -0.02em;
      text-align: right;
      max-width: 100%;
    }

    .metric-row-text {
      grid-template-columns: 14px minmax(0, 1fr);
      align-items: start;
    }

    .metric-row-text .metric-value {
      grid-column: 2;
      white-space: normal;
      text-align: left;
      line-height: 1.6;
      font-weight: 700;
      overflow-wrap: anywhere;
      padding-top: 2px;
    }

    .metric-dot {
      width: 10px;
      height: 10px;
      border-radius: 50%;
      display: inline-block;
      box-shadow: 0 0 0 5px transparent;
    }

    .metric-dot.red { background: var(--red); box-shadow: 0 0 0 5px rgba(178, 58, 49, 0.12); }
    .metric-dot.amber { background: var(--amber); box-shadow: 0 0 0 5px rgba(166, 106, 23, 0.12); }
    .metric-dot.green { background: var(--green); box-shadow: 0 0 0 5px rgba(38, 119, 90, 0.12); }
    .metric-dot.info { background: var(--info); box-shadow: 0 0 0 5px rgba(55, 100, 141, 0.12); }

    .section-insight {
      margin-top: 14px;
      padding: 14px 15px;
      border-radius: 18px;
      border: 1px solid rgba(80, 112, 146, 0.18);
      background: linear-gradient(180deg, rgba(249, 252, 255, 0.98) 0%, rgba(241, 247, 253, 0.92) 100%);
      box-shadow: inset 0 1px 0 rgba(255, 255, 255, 0.7);
      font-size: 13px;
      line-height: 1.7;
      color: var(--muted-strong);
    }

    .narrative-header {
      display: flex;
      align-items: center;
      gap: 10px;
      margin-bottom: 8px;
      flex-wrap: wrap;
    }

    .section-insight strong {
      display: inline-block;
      margin: 0;
      font-size: 11px;
      letter-spacing: 0.12em;
      text-transform: uppercase;
      color: var(--muted);
      font-weight: 800;
    }

    .ai-badge {
      display: inline-flex;
      align-items: center;
      gap: 6px;
      padding: 4px 8px;
      border-radius: 999px;
      border: 1px solid rgba(80, 112, 146, 0.2);
      background: rgba(255, 255, 255, 0.7);
      color: var(--info);
      font-size: 11px;
      font-weight: 800;
      letter-spacing: 0.08em;
      text-transform: uppercase;
      white-space: nowrap;
    }

    .ai-badge svg {
      width: 12px;
      height: 12px;
      flex: 0 0 auto;
      stroke: currentColor;
      fill: none;
      stroke-width: 1.6;
      stroke-linecap: round;
      stroke-linejoin: round;
    }

    .mode-badge {
      display: inline-flex;
      align-items: center;
      gap: 6px;
      padding: 4px 8px;
      border-radius: 999px;
      border: 1px solid rgba(19, 35, 56, 0.14);
      background: rgba(255, 255, 255, 0.75);
      color: var(--text-strong);
      font-size: 11px;
      font-weight: 800;
      letter-spacing: 0.08em;
      text-transform: uppercase;
      white-space: nowrap;
    }

    .decision-stack {
      display: grid;
      gap: var(--space-4);
    }

    .decision-card { padding-bottom: var(--space-4); }

    .decision-card.insights {
      background:
        radial-gradient(circle at top right, rgba(55, 100, 141, 0.08), rgba(55, 100, 141, 0) 36%),
        linear-gradient(180deg, rgba(252, 253, 255, 0.98) 0%, rgba(245, 249, 253, 0.96) 100%);
    }

    .decision-card.actions {
      background:
        radial-gradient(circle at top right, rgba(166, 106, 23, 0.09), rgba(166, 106, 23, 0) 36%),
        linear-gradient(180deg, rgba(255, 254, 251, 0.98) 0%, rgba(252, 247, 239, 0.96) 100%);
    }

    .decision-card.risks {
      background:
        radial-gradient(circle at top right, rgba(178, 58, 49, 0.09), rgba(178, 58, 49, 0) 36%),
        linear-gradient(180deg, rgba(255, 252, 252, 0.98) 0%, rgba(252, 243, 242, 0.96) 100%);
    }

    .decision-list {
      list-style: none;
      margin: 0;
      padding: var(--space-4) var(--space-5) 0;
      display: grid;
      gap: 12px;
    }

    .decision-list li {
      position: relative;
      padding: 16px 18px 16px 20px;
      border-radius: 18px;
      border: 1px solid rgba(19, 35, 56, 0.08);
      background: rgba(255, 255, 255, 0.72);
      font-size: 14px;
      line-height: 1.7;
      color: var(--text);
      box-shadow: inset 0 1px 0 rgba(255, 255, 255, 0.6);
    }

    .decision-list li::before {
      content: "";
      position: absolute;
      left: 0;
      top: 12px;
      bottom: 12px;
      width: 3px;
      border-radius: 999px;
      background: var(--accent);
    }

    .decision-card.actions .decision-list li::before { background: var(--amber); }
    .decision-card.risks .decision-list li::before { background: var(--red); }

    .ai-footnote {
      margin: 14px var(--space-5) 0;
      padding-top: 12px;
      border-top: 1px solid rgba(19, 35, 56, 0.08);
      display: flex;
      align-items: center;
      gap: 10px;
      font-size: 11px;
      line-height: 1.5;
      color: var(--muted);
      flex-wrap: wrap;
    }

    .ai-footnote .ai-badge { flex: 0 0 auto; }

    .table-panel { position: relative; }

    .table-wrap {
      padding: 0 var(--space-4) var(--space-4);
      overflow-x: auto;
      background: linear-gradient(180deg, rgba(248, 250, 252, 0.58) 0%, rgba(255, 255, 255, 0) 100%);
    }

    table {
      width: 100%;
      min-width: 920px;
      border-collapse: separate;
      border-spacing: 0;
      font-size: 13px;
      color: var(--text);
    }

    th, td {
      text-align: left;
      padding: 13px 14px;
      vertical-align: middle;
      border-bottom: 1px solid rgba(19, 35, 56, 0.08);
    }

    th {
      position: sticky;
      top: 0;
      z-index: 2;
      font-size: 10px;
      text-transform: uppercase;
      letter-spacing: 0.14em;
      color: var(--muted);
      background: rgba(242, 246, 250, 0.96);
      backdrop-filter: blur(8px);
      font-weight: 800;
      box-shadow: inset 0 -1px 0 rgba(19, 35, 56, 0.08);
    }

    tbody tr:nth-child(even) td { background: rgba(248, 250, 252, 0.62); }
    tbody tr:hover td { background: rgba(239, 245, 250, 0.95); }
    tbody tr:last-child td { border-bottom: none; }
    td:first-child { font-weight: 700; color: var(--muted-strong); }
    td.value { font-weight: 800; color: var(--text-strong); }
    td.value, td.target, td.status-cell { white-space: nowrap; }

    .drilldown-stack {
      display: grid;
      gap: var(--space-4);
    }

    .drilldown-card {
      margin: 0 var(--space-4) var(--space-4);
      padding: 18px;
      border-radius: 20px;
      border: 1px solid rgba(19, 35, 56, 0.08);
      background: rgba(250, 252, 255, 0.88);
      scroll-margin-top: 18px;
    }

    .drilldown-card summary {
      cursor: pointer;
      font-weight: 800;
      color: var(--text-strong);
      list-style: none;
      display: flex;
      align-items: center;
      gap: 10px;
    }

    .drilldown-card summary::-webkit-details-marker { display: none; }
    .drilldown-card summary::before {
      content: "▾";
      font-size: 14px;
      color: var(--accent);
      line-height: 1;
    }
    .drilldown-card:not([open]) summary::before { content: "▸"; }

    .drilldown-meta {
      margin: 10px 0 14px;
      font-size: 12px;
      line-height: 1.6;
      color: var(--muted);
    }

    .drilldown-table {
      width: 100%;
      min-width: 720px;
      border-collapse: separate;
      border-spacing: 0;
      font-size: 12px;
    }

    .drilldown-empty {
      font-size: 12px;
      color: var(--muted);
      padding-top: 10px;
    }

    .raw-detail-title {
      margin: 16px 0 8px;
      font-size: 11px;
      font-weight: 800;
      letter-spacing: 0.1em;
      text-transform: uppercase;
      color: var(--muted);
    }

    .group-detail-list {
      display: grid;
      gap: 10px;
      margin-top: 14px;
    }

    .group-detail-card {
      border: 1px solid rgba(19, 35, 56, 0.08);
      border-radius: 16px;
      background: rgba(255, 255, 255, 0.72);
      overflow: hidden;
    }

    .group-detail-card summary {
      cursor: pointer;
      padding: 12px 14px;
      font-size: 12px;
      font-weight: 700;
      color: var(--text-strong);
      list-style: none;
      display: flex;
      align-items: center;
      gap: 8px;
    }

    .group-detail-card summary::-webkit-details-marker { display: none; }
    .group-detail-card summary::before {
      content: "▾";
      font-size: 13px;
      color: var(--accent);
      line-height: 1;
    }
    .group-detail-card:not([open]) summary::before { content: "▸"; }

    .footer-note {
      margin-top: 12px;
      display: flex;
      justify-content: space-between;
      gap: 16px;
      align-items: center;
      flex-wrap: wrap;
      font-size: 12px;
      line-height: 1.5;
      color: var(--muted);
    }

    .footer-note strong { color: var(--muted-strong); }

    @media (max-width: 1320px) {
      .summary-grid { grid-template-columns: repeat(3, minmax(0, 1fr)); }
      .main-grid { grid-template-columns: 1fr; }
    }

    @media (max-width: 960px) {
      .page { width: min(100% - 28px, 1880px); margin-top: 16px; gap: var(--space-4); }
      .hero { padding: var(--space-5); }
      .hero-top { grid-template-columns: 1fr; }
      .hero-metadata { grid-template-columns: repeat(2, minmax(0, 1fr)); }
      .summary-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); }
      .section-grid { grid-template-columns: 1fr; }
    }

    @media (max-width: 640px) {
      .page { width: calc(100% - 16px); }
      .hero, .panel-header, .decision-list, .table-wrap { padding-left: var(--space-3); padding-right: var(--space-3); }
      .hero { padding-top: var(--space-4); padding-bottom: var(--space-4); }
      .hero-metadata, .summary-grid { grid-template-columns: 1fr; }
      .meta-tile, .summary-card, .section-card { border-radius: 18px; }
      .section-card { min-height: auto; }
      .metric-row { grid-template-columns: 14px minmax(0, 1fr); }
      .metric-value { grid-column: 2; }
      .ai-footnote, .footer-note { align-items: flex-start; flex-direction: column; }
    }
  </style>
</head>
<body>
  <div class="page">
    <section class="hero">
      <div class="hero-top">
        <div class="hero-copy">
          <div class="eyebrow">KPI Intelligence Agent</div>
          <h1>{{ payload.header.title }}</h1>
          <div class="subtitle">{{ payload.header.subtitle }}</div>        </div>
        <div class="status-badge {{ payload.header.overall_status }}">{{ payload.header.overall_status.upper() }} Overall Status</div>
      </div>
      <div class="hero-metadata">
        <div class="meta-tile">
          <div class="meta-label">Reporting Period</div>
          <div class="meta-value">
            {{ payload.reporting_period.label }}<br />
            {{ payload.reporting_period.start_date }} to {{ payload.reporting_period.end_date }}
          </div>
        </div>
        <div class="meta-tile">
          <div class="meta-label">Warehouse Scope</div>
          <div class="meta-value">{{ payload.header.warehouse_scope | join(', ') }}</div>
        </div>
        <div class="meta-tile">
          <div class="meta-label">SKU Family Scope</div>
          <div class="meta-value">{{ payload.header.sku_family_scope | join(', ') }}</div>
        </div>
        <div class="meta-tile">
          <div class="meta-label">Calculation Version</div>
          <div class="meta-value">{{ payload.header.calculation_version }}</div>
        </div>
        <div class="meta-tile">
          <div class="meta-label">Generated At</div>
          <div class="meta-value">{{ payload.header.generated_at }}</div>
        </div>
      </div>
    </section>

    <section class="summary-grid">
      {% for card in payload.summary_cards %}
      <article class="summary-card {{ card.status }}">
        <div class="card-top">
          <div class="card-label">{{ card.label }}</div>
          <div class="pill {{ card.status }}">{{ card.status.upper() }}</div>
        </div>
        <div class="card-value">{{ card.display_value }}</div>
        <div class="card-footer">
          <div class="card-target"><strong>Target:</strong> {{ card.target_display or 'n/a' }}</div>
          {% if card.note %}
          <div class="card-note"><strong>Note:</strong> {{ card.note }}</div>
          {% endif %}
          {% if card.drilldown_anchor %}
          <a class="drilldown-link" href="#{{ card.drilldown_anchor }}">View Drill-Down</a>
          {% endif %}
        </div>
      </article>
      {% endfor %}
    </section>

    <section class="main-grid">
      <div class="panel">
        <div class="panel-header">
          <div>
            <h2 class="panel-title">Operational Domains</h2>
            <div class="panel-subtitle">Structured operational view across inbound, outbound, inventory, warehouse, and workforce performance for rapid executive scanning.</div>
          </div>
        </div>
        <div class="section-grid">
          {% for section in payload.sections %}
          <article class="section-card {{ section.risk_level }}">
            <div class="section-top">
              <h3 class="section-title">{{ section.name }}</h3>
              <div class="pill {{ section.risk_level }}">{{ section.risk_level.upper() }} Risk</div>
            </div>
            <ul class="metric-list">
              {% for kpi in section.kpis %}
              <li class="metric-row{% if kpi.unit == 'text' %} metric-row-text{% endif %}">
                <span class="metric-dot {{ kpi.status }}"></span>
                <span class="metric-name">
                  {{ kpi.name }}
                  {% if kpi.drilldown_anchor %}
                  <br /><a class="drilldown-jump" href="#{{ kpi.drilldown_anchor }}">Open Drill-Down</a>
                  {% endif %}
                </span>
                <span class="metric-value">{{ kpi.display_value }}</span>
              </li>
              {% endfor %}
            </ul>
            <div class="section-insight">
              <div class="narrative-header">
                <strong>Section Insight</strong>
                {% if payload.metadata.narrative_generation == 'llm_refined' %}
                <span class="ai-badge">
                  <svg viewBox="0 0 16 16" aria-hidden="true">
                    <path d="M8 1.5 9.6 6.4 14.5 8l-4.9 1.6L8 14.5 6.4 9.6 1.5 8l4.9-1.6z"></path>
                  </svg>
                  AI
                </span>
                {% else %}
                <span class="mode-badge">Deterministic</span>
                {% endif %}
              </div>
              {{ section.insight }}
            </div>
          </article>
          {% endfor %}
        </div>
      </div>

      <div class="decision-stack">
        <section class="panel decision-card insights">
          <div class="panel-header">
            <div>
              <div class="narrative-header">
                <h2 class="panel-title">Key Insights</h2>
                {% if payload.metadata.narrative_generation == 'llm_refined' %}
                <span class="ai-badge">
                  <svg viewBox="0 0 16 16" aria-hidden="true">
                    <path d="M8 1.5 9.6 6.4 14.5 8l-4.9 1.6L8 14.5 6.4 9.6 1.5 8l4.9-1.6z"></path>
                  </svg>
                  AI
                </span>
                {% else %}
                <span class="mode-badge">Deterministic</span>
                {% endif %}
              </div>
              <div class="panel-subtitle">Leadership-level signals that explain current operational pressure.</div>
            </div>
          </div>
          <ul class="decision-list">
            {% for item in payload.insights %}
            <li>{{ item }}</li>
            {% endfor %}
          </ul>
        </section>

        <section class="panel decision-card actions">
          <div class="panel-header">
            <div>
              <div class="narrative-header">
                <h2 class="panel-title">Recommended Actions</h2>
                {% if payload.metadata.narrative_generation == 'llm_refined' %}
                <span class="ai-badge">
                  <svg viewBox="0 0 16 16" aria-hidden="true">
                    <path d="M8 1.5 9.6 6.4 14.5 8l-4.9 1.6L8 14.5 6.4 9.6 1.5 8l4.9-1.6z"></path>
                  </svg>
                  AI
                </span>
                {% else %}
                <span class="mode-badge">Deterministic</span>
                {% endif %}
              </div>
              <div class="panel-subtitle">Priority actions for supply, inventory, warehouse, and workforce recovery.</div>
            </div>
          </div>
          <ul class="decision-list">
            {% for item in payload.recommendations %}
            <li>{{ item }}</li>
            {% endfor %}
          </ul>
       
        </section>
      </div>
    </section>

    <section class="table-panel">
      <div class="panel-header">
        <div>
          <h2 class="panel-title">KPI Summary Table</h2>
          <div class="panel-subtitle">Compact reference table for current value, target, and health status with improved scanability for leadership review.</div>
        </div>
      </div>
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>KPI</th>
              <th>Domain</th>
              <th>Value</th>
              <th>Unit</th>
              <th>Target</th>
              <th>Delta</th>
              <th>Status</th>
              <th>Comment</th>
            </tr>
          </thead>
          <tbody>
            {% for row in payload.kpi_table %}
            <tr>
              <td>{{ row.kpi }}</td>
              <td>{{ row.domain }}</td>
              <td class="value">{{ row.display_value }}</td>
              <td>{{ row.unit }}</td>
              <td class="target">{{ row.target_display or 'n/a' }}</td>
              <td class="value">
                {% if row.target is not none %}
                {{ '%+.4f'|format(row.value - row.target) }}
                {% else %}
                n/a
                {% endif %}
              </td>
              <td class="status-cell"><span class="pill {{ row.status }}">{{ row.status.upper() }}</span></td>
              <td>
                {{ row.trend_note or 'Current-period view' }}
                {% if row.drilldown_anchor %}
                <br /><a class="drilldown-jump" href="#{{ row.drilldown_anchor }}">Open Drill-Down</a>
                {% endif %}
              </td>
            </tr>
            {% endfor %}
          </tbody>
        </table>
        <div class="footer-note">
          <span><strong>Generated by:</strong> KPI Intelligence Agent</span>
          <span><strong>Reporting window:</strong> {{ payload.reporting_period.start_date }} to {{ payload.reporting_period.end_date }}</span>
        </div>
      </div>
    </section>

    <section class="table-panel">
      <div class="panel-header">
        <div>
          <h2 class="panel-title">Drill-Down Detail</h2>
          <div class="panel-subtitle">Each KPI opens its own deterministic detail table, filtered to the specific metric that was clicked.</div>
        </div>
      </div>
      <div class="drilldown-stack">
        {% for section in payload.sections %}
          {% for drilldown in section.html_kpi_drilldowns %}
          <details class="drilldown-card" id="{{ drilldown.anchor }}">
            <summary>{{ drilldown.section_name }} | {{ drilldown.kpi_name }}</summary>
            <div class="drilldown-meta">
              Detail view: {{ drilldown.label }} |
              Source: {{ drilldown.source_dataset }} |
              Warehouse filter: {{ drilldown.applied_filters.warehouse_filter or 'All' }} |
              SKU family filter: {{ drilldown.applied_filters.sku_family_filter or 'All' }}
            </div>
            {% if drilldown.logic_note %}
            <div class="drilldown-meta">{{ drilldown.logic_note }}</div>
            {% endif %}
            {% if not drilldown.available or not drilldown.rows %}
            <div class="drilldown-empty">{{ drilldown.unavailable_reason }}</div>
            {% else %}
            <div class="table-wrap" style="padding:0;">
              <table class="drilldown-table">
                <thead>
                  <tr>
                    {% if drilldown.rows[0].rank is not none %}
                    <th>Rank</th>
                    {% endif %}
                    {% for column in drilldown.group_by %}
                    <th>{{ drilldown.dimension_labels[column] }}</th>
                    {% endfor %}
                    <th>{{ drilldown.kpi_name }}</th>
                    <th>Status</th>
                    <th>Formula</th>
                  </tr>
                </thead>
                <tbody>
                  {% for row in drilldown.rows %}
                  <tr>
                    {% if row.rank is not none %}
                    <td>{{ row.rank }}</td>
                    {% endif %}
                    {% for column in drilldown.group_by %}
                    <td>{{ row.dimensions[column] }}</td>
                    {% endfor %}
                    <td>{{ row.metric.display_value }}</td>
                    <td>{% if row.metric.status and row.metric.status != 'info' %}{{ row.metric.status.upper() }}{% else %}n/a{% endif %}</td>
                    <td>{{ row.metric.formula }}</td>
                  </tr>
                  {% endfor %}
                </tbody>
              </table>
            </div>
            <div class="raw-detail-title">Group-Specific Raw Records</div>
            <div class="group-detail-list">
              {% for row in drilldown.rows %}
              <details class="group-detail-card" id="{{ row.group_anchor }}">
                <summary>
                  {% if row.rank is not none %}Rank {{ row.rank }} | {% endif %}
                  {% for column in drilldown.group_by %}
                  {{ drilldown.dimension_labels[column] }}: {{ row.dimensions[column] }}{% if not loop.last %} | {% endif %}
                  {% endfor %}
                  | {{ row.metric.display_value }}
                </summary>
                {% if row.raw_detail and row.raw_detail.rows %}
                <div class="drilldown-meta" style="padding: 0 14px 12px;">{{ row.raw_detail.logic_note }} Row count: {{ row.raw_detail.row_count }}</div>
                <div class="table-wrap" style="padding:0 14px 14px;">
                  <table class="drilldown-table">
                    <thead>
                      <tr>
                        {% for column in row.raw_detail.columns %}
                        <th>{{ column }}</th>
                        {% endfor %}
                      </tr>
                    </thead>
                    <tbody>
                      {% for raw_row in row.raw_detail.rows %}
                      <tr>
                        {% for column in row.raw_detail.columns %}
                        <td>{{ raw_row[column] }}</td>
                        {% endfor %}
                      </tr>
                      {% endfor %}
                    </tbody>
                  </table>
                </div>
                {% else %}
                <div class="drilldown-empty" style="padding: 0 14px 14px;">No raw records matched this grouped row.</div>
                {% endif %}
              </details>
              {% endfor %}
            </div>
            {% endif %}
          </details>
          {% endfor %}
        {% endfor %}
      </div>
    </section>
  </div>
  <script>
    (function () {
      function openTargetFromHash() {
        if (!window.location.hash) return;
        var target = document.getElementById(window.location.hash.slice(1));
        if (!target) return;
        var details = target.closest('details');
        if (details) details.open = true;
      }
      window.addEventListener('hashchange', openTargetFromHash);
      openTargetFromHash();
    }());
  </script>
</body>
</html>
"""



def write_html(payload: Dict, output_dir: Path) -> Path:
    # ================================
    # Function: write_html
    # Purpose: Renders the leadership one-pager as a standalone HTML file.
    # Inputs:
    #   - payload (Dict): packaged KPI output
    #   - output_dir (Path): target directory
    # Output:
    #   - Path to the written HTML file
    # ================================
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / "one_pager.html"
    html = Template(HTML_TEMPLATE).render(payload=payload)
    path.write_text(html, encoding="utf-8")
    return path


def write_run_summary(payload: Dict, output_dir: Path) -> Path:
    # ================================
    # Function: write_run_summary
    # Purpose: Writes a compact text summary for quick inspection and audit.
    # Inputs:
    #   - payload (Dict): packaged KPI output
    #   - output_dir (Path): target directory
    # Output:
    #   - Path to the written text file
    # ================================
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / "run_summary.txt"
    lines = [
        payload["header"]["title"],
        f"Reporting period: {payload['reporting_period']['label']} ({payload['reporting_period']['start_date']} to {payload['reporting_period']['end_date']})",
        f"Overall status: {payload['header']['overall_status'].upper()}",
        f"Warehouse scope: {', '.join(payload['header']['warehouse_scope'])}",
        f"SKU family scope: {', '.join(payload['header']['sku_family_scope'])}",
        "",
        "Summary cards:",
    ]
    for card in payload["summary_cards"]:
        lines.append(f"- {card['label']}: {card['display_value']} [{card['status'].upper()}] target={card['target_display'] or 'n/a'}")
    lines.append("")
    lines.append("Top risks:")
    lines.extend(f"- {risk}" for risk in payload.get("risks", []))
    lines.append("")
    lines.append("Validation warnings:")
    if payload["audit"]["validation_warnings"]:
        lines.extend(f"- {w}" for w in payload["audit"]["validation_warnings"])
    else:
        lines.append("- none")

    path.write_text("\n".join(lines), encoding="utf-8")
    return path
